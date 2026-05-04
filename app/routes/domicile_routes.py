import re
import base64
import logging
from typing import Optional
from tkinter import messagebox
from bs4 import BeautifulSoup
from fastapi import APIRouter, HTTPException, Body, Query, File, UploadFile
from io import BytesIO
from openpyxl import load_workbook

from app import config
from app.database import open_con
from app.nitb import approve, nitb_session, get_session
from datetime import date, datetime
router = APIRouter()
logger = logging.getLogger(__name__)


@router.post("/approve/{id}")
def approve_request(id: int):
    result = approve(f"{config.NITB_BASE}/domicile/application/{id}", "domicile-approval")
    if result is None or not result.get("success", True):
        logger.error("NITB approval failed for id %s: %s", id, result)
        raise HTTPException(status_code=500, detail="NITB approval failed")
    return {"status": "success", "message": f"Request {id} approved"}


@router.get("/check/{cnic}")
def check_domicile_status(cnic: str):
    if not cnic:
        raise HTTPException(status_code=400, detail="CNIC is required")

    # Remove non-digits (handles both '61101-4561237-8' and '6110145612378')
    clean_cnic = re.sub(r"\D", "", cnic)
    if len(clean_cnic) != 13:
        raise HTTPException(status_code=400, detail="CNIC must be 13 digits (e.g., 61101-4561237-8)")

    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        cur.execute(
            "SELECT receipt_no, Status, First_Name, remarks FROM domicile WHERE cnic = %s",
            (clean_cnic,)
        )
        result = cur.fetchone()
        cur.close()
        con.close()

        if result:
            return {"status": result}
        else:
            raise HTTPException(status_code=404, detail="Record not found")

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error checking domicile status for CNIC %s", clean_cnic)
        raise HTTPException(status_code=500, detail=f"Something went wrong: {str(exc)}")


@router.get('/statistics/check')
def statistics():
    global nitb_session
    if nitb_session is None:
        nitb_session = get_session()
        if not nitb_session:
            raise HTTPException(status_code=500, detail="Unable to connect to NITB")

    url = f"{config.NITB_BASE}/dashboard/statistics"
    try:
        page = nitb_session.get(url, timeout=30)
        logger.debug("Statistics page fetched, length=%s", len(page.content))
        page.raise_for_status()
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to fetch IDP data")

    soup = BeautifulSoup(page.content, 'html.parser')
    details = {}

    for divs in soup.find_all('div', class_='bd-highlight'):
        data = divs.text.split()
        if not data:
            continue
        if data[0].strip() == 'Domicile' and len(data) >= 4:
            details['domicile'] = int(data[3]) if data[3].isdigit() else data[3]
        elif data[0].strip() == 'IDP' and len(data) >= 4:
            details['idp'] = int(data[3]) if data[3].isdigit() else data[3]

    if details:
        return details

    raise HTTPException(status_code=404, detail="No record found or IDP not approved")

@router.post("/approve/{id}")
def approve_request(id: int):
    
    result = approve(f"https://admin-icta.nitb.gov.pk/domicile/application/{id}", "domicile-approval")
    if result is None or not result.get("success", True):
        raise HTTPException(status_code=500, detail="NITB approval failed")
    return {"status": "success", "message": f"Request {id} approved"}

@router.get("/check/{cnic}")
def check_domicile_status(cnic: str):
    if not cnic:
        return {"error": "CNIC is required"}

    # Remove non-digits (handles both '61101-4561237-8' and '6110145612378')
    clean_cnic = re.sub(r"\D", "", cnic)
    if len(clean_cnic) != 13:
        return {"error": "CNIC must be 13 digits (e.g., 61101-4561237-8)"}

    
    try:
        con, cur = open_con()
        if type(con) is str:
            return {"error": f"Database connection failed: {cur}"}

        cur.execute(
            "SELECT receipt_no, Status, First_Name, remarks FROM domicile WHERE cnic = %s",
            (clean_cnic,)
        )
        result = cur.fetchone()
        cur.close()
        con.close()

        if result:
            return {"status": result}
        else:
            return {"error": "Record not found"}

    except Exception as exc:
        return {"error": f"Something went wrong: {str(exc)}"}

@router.get('/statistics/check')
def statistics():
    global nitb_session
    if nitb_session is None:
        nitb_session = get_session()
        if not nitb_session:
            raise HTTPException(status_code=500, detail="Unable to connect to NITB")

    url = f"{config.NITB_BASE}/dashboard/statistics"
    try:
        page = nitb_session.get(url, timeout=30)
        print(page.content)
        page.raise_for_status()
    except Exception:
        raise HTTPException(status_code=500, detail="Failed to fetch IDP data")

    soup = BeautifulSoup(page.content, 'html.parser')
    details = {}

    for divs in soup.find_all('div', class_='bd-highlight'):
        data = divs.text.split()
        if not data:
            continue
        if data[0].strip() == 'Domicile' and len(data) >= 4:
            details['domicile'] = int(data[3]) if data[3].isdigit() else data[3]
        elif data[0].strip() == 'IDP' and len(data) >= 4:
            details['idp'] = int(data[3]) if data[3].isdigit() else data[3]

    if details:
        return details  # 👈 return dict directly instead of {'result': details}

    raise HTTPException(status_code=404, detail="No record found or IDP not approved")


@router.get('/approvers')
def get_approvers():
    """Return all rows from the `approvers` table as JSON."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        cur.execute("SELECT * FROM approvers;")
        rows = cur.fetchall()
        cur.close()
        con.close()

        return {"approvers": rows}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch approvers")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/tehsils')
def get_tehsils(district_id: Optional[int] = Query(None, description="Filter tehsils by district_id")):
    """Return rows from the `Tehsils` table as JSON. If `district_id` is provided,
    return tehsils for that district only (useful for dependent dropdowns)."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        if district_id is not None:
            cur.execute("SELECT * FROM Tehsils WHERE district_id = %s;", [district_id])
        else:
            cur.execute("SELECT * FROM Tehsils;")

        rows = cur.fetchall()
        cur.close()
        con.close()

        return {"tehsils": rows}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch tehsils")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/districts')
def get_districts():
    """Return all rows from the `Districts` table as JSON."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        cur.execute("Select * from Districts;")
        rows = cur.fetchall()
        cur.close()
        con.close()

        return {"districts": rows}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch districts")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/black-list')
def get_black_list():
    """Return all black listed CNICs."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        Query = """Select b.black_list_id,
                          b.cnic,
                          b.status,
                          b.reason,
                          b.clearance_reason,
                          u.user_name,
                          b.created_at 
                from black_list as b 
                Join users as u
                    on b.user_id = u.user_id
                order by black_list_id Desc Limit 200;"""
        cur.execute(Query)
        rows = cur.fetchall()
        cur.close()
        con.close()

        return {"black_list": rows}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch black list")
        raise HTTPException(status_code=500, detail=str(exc))


@router.post('/black-list')
def add_to_black_list(cnic: str = Query(...), reason: str = Query(...), user_id: int = Query(...)):
    """Add a CNIC to the black list."""
    try:
        if not cnic or len(cnic) != 13:
            raise HTTPException(status_code=400, detail="CNIC must be 13 digits")
        
        if not cnic.isnumeric():
            raise HTTPException(status_code=400, detail="CNIC must contain only digits")
        
        if not reason or len(reason) < 10:
            raise HTTPException(status_code=400, detail="Reason must be at least 10 characters")

        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        # Check if CNIC already exists
        cur.execute("Select cnic from black_list where cnic = %s;", [cnic])
        existing = cur.fetchone()

        if existing:
            # Update existing record
            cur.execute("Update black_list set reason=%s, status='Blocked' where cnic=%s;", [reason, cnic])
            con.commit()
            cur.execute("Insert Into black_list_history (remarks, user_id) values (%s, %s);", 
                       [f"Existing Record Updated. CNIC Blocked. {reason}", user_id])
            con.commit()
            cur.close()
            con.close()
            return {"status": "success", "message": "CNIC record updated and blocked"}
        else:
            # Insert new record
            cur.execute("Insert Into black_list (cnic, reason, user_id) values (%s, %s, %s);", 
                       [cnic, reason, user_id])
            con.commit()
            cur.execute("select Last_insert_id();")
            last_id = cur.fetchone()
            cur.execute("Insert Into black_list_history (black_list_id, remarks, user_id) values (%s, %s, %s);",
                       [last_id[0], f"CNIC Blocked. {reason}", user_id])
            con.commit()
            cur.close()
            con.close()
            return {"status": "success", "message": "CNIC added to black list"}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to add CNIC to black list")
        raise HTTPException(status_code=500, detail=str(exc))


@router.delete('/black-list/{black_list_id}')
def remove_from_black_list(black_list_id: int, clearance_reason: str = Query(...), user_id: int = Query(...)):
    """Remove a CNIC from the black list (unblock)."""
    try:
        if not clearance_reason or len(clearance_reason) < 10:
            raise HTTPException(status_code=400, detail="Clearance reason must be at least 10 characters")

        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        cur.execute("Update black_list set clearance_reason=%s, status='Unblocked' where black_list_id=%s;",
                   [clearance_reason, black_list_id])
        con.commit()
        cur.execute("Insert Into black_list_history (black_list_id, remarks, user_id) values (%s, %s, %s);",
                   [black_list_id, f"CNIC Unblocked. {clearance_reason}", user_id])
        con.commit()
        cur.close()
        con.close()

        return {"status": "success", "message": "CNIC removed from black list"}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to remove CNIC from black list")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/childern')
def get_childern(father_cnic: str = Query(..., description="Father's CNIC number")):
    """Get all children records for a given father CNIC."""
    try:
        if not father_cnic or len(father_cnic.strip()) == 0:
            raise HTTPException(status_code=400, detail="father_cnic is required")

        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        Query = "Select * from childern where father_cnic = %s order by child_dob asc;"
        cur.execute(Query, [father_cnic])
        rows = cur.fetchall()
        cur.close()
        con.close()

        if rows:
            print(rows)
            # Format children records for response
            formatted_children = []
            for row in rows:
                formatted_children.append({
                    "child_id": row.get('child_id'),
                    "father_cnic": row['Father_CNIC'],
                    "child_name": row['Child_Name'],
                    "child_dob": str(row['Child_dob']) if row['Child_dob'] else None,
                    "child_gender": row.get('child_gender'),
                })
            return {"childern": formatted_children}
        else:
            return {"childern": []}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch children for father_cnic %s", father_cnic)


@router.get('/reports/summary')
def reports_summary(typ: str = Query(..., description="Report type"),
                    from_date: str = Query(..., description="From date YYYY-MM-DD"),
                    to_date: str = Query(..., description="To date YYYY-MM-DD"),
                    officer_name: str | None = Query(None, description="Approver name for Approved Files")):
    """Return summary counts for requested report type between two dates."""
    # validate dates
    try:
        datetime.strptime(from_date, '%Y-%m-%d')
        datetime.strptime(to_date, '%Y-%m-%d')
    except Exception:
        raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DD format")

    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        if typ == 'Issued Domicile':
            Query = "Select Count(Govt_Fee) as Total_Applications from cash_report Where Duplicate_Entry = 'No' And Application_Type = 'offline' And Request_Type= 'New' And Domicile_Date Between %s and %s;"
            parm_list = [from_date, to_date]
        elif typ == 'Accepted Files':
            Query = """SELECT count(dom_id) as no_of_domiciles 
                    FROM domicile_reports.domicile 
                    Where Dom_Date Between %s and %s;"""
            parm_list = [from_date, to_date]
        elif typ == 'Verification Letters':
            Query = "select count(letter_id) as issued_letters from verification_letters where timestamp between %s and %s;"
            parm_list = [from_date, to_date]
        elif typ == 'Approved Files':
            if not officer_name:
                raise HTTPException(status_code=400, detail="officer_name is required for Approved Files report")
            Query = "select count(dom_id) as approved_files from domicile where approver_id = (select approver_id from approvers where approver_name = %s) and Dom_Date Between %s and %s;"
            parm_list = [officer_name, from_date, to_date]
        else:
            raise HTTPException(status_code=400, detail="Unknown report type")

        cur.execute(Query, parm_list)
        count_data = cur.fetchall()
        cur.close()
        con.close()

        # additional computation for issued domicile
        extra = {}
        if typ == 'Issued Domicile' and count_data:
            try:
                total_apps = int(count_data[0].get('Total_Applications') or 0)
                extra['total_amount'] = total_apps * 200
            except Exception:
                pass

        return {"report_type": typ, "from": from_date, "to": to_date, "result": count_data, "extra": extra}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to generate summary report")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/reports/daily')
def reports_daily(typ: str = Query(..., description="Report type"),
                  from_date: str = Query(..., description="From date YYYY-MM-DD"),
                  to_date: str = Query(..., description="To date YYYY-MM-DD"),
                  officer_name: str | None = Query(None, description="Approver name for Approved Files")):
    """Return per-day counts for requested report type between two dates."""
    try:
        datetime.strptime(from_date, '%Y-%m-%d')
        datetime.strptime(to_date, '%Y-%m-%d')
    except Exception:
        raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DD format")

    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        if typ == 'Issued Domicile':
            Query = """SELECT Domicile_Date, count(Domicile_ID) as no_of_domiciles 
                    FROM domicile_reports.cash_report 
                    Where Duplicate_Entry = 'No'
                    and Application_Type = 'offline'
                    and Domicile_Date Between %s and %s group by domicile_date;"""
            parm_list = [from_date, to_date]
        elif typ == 'Accepted Files':
            Query = """SELECT Dom_Date, count(dom_id) as no_of_domiciles 
                    FROM domicile_reports.domicile 
                    Where Process_Type <> 'Urgent'
                    and Dom_Date Between %s and %s group by Dom_Date order by Dom_Date;"""
            parm_list = [from_date, to_date]
        elif typ == 'Verification Letters':
            Query = "select count(letter_id) as Total_Letters, date(timestamp) from verification_letters where timestamp between %s and %s group by Date(timestamp);"
            parm_list = [from_date, to_date]
        elif typ == 'Approved Files':
            if not officer_name:
                raise HTTPException(status_code=400, detail="officer_name is required for Approved Files report")
            Query = "select count(dom_id) as Approved_Files, dom_date from domicile where approver_id = (select approver_id from approvers where approver_name = %s) and dom_date between %s and %s group by dom_date;"
            parm_list = [officer_name, from_date, to_date]
        else:
            raise HTTPException(status_code=400, detail="Unknown report type")

        cur.execute(Query, parm_list)
        rows = cur.fetchall()

        approver_data = None
        if typ == 'Approved Files':
            cur.execute("Select * from approvers where approver_name = %s;", [officer_name])
            approver_data = cur.fetchall()

        cur.close()
        con.close()

        return {"report_type": typ, "from": from_date, "to": to_date, "rows": rows, "approver": approver_data}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to generate daily report")
        raise HTTPException(status_code=500, detail=str(exc))
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/black-list/history/{black_list_id}')
def get_black_list_history(black_list_id: int):
    """Get the history of a black listed CNIC."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        Query = "Select h.remarks, h.created_at, u.user_name from black_list_history as h Join users as u on u.user_id = h.user_id where h.black_list_id = %s;"
        cur.execute(Query, [black_list_id])
        rows = cur.fetchall()
        cur.close()
        con.close()

        if rows:
            return {"history": rows}
        else:
            return {"history": [], "message": "No history found"}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch black list history")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/records/search')
def search_records(search_type: str = Query(None), search_input: str = Query(None)):
    """
    Search domicile records based on various criteria.
    
    Search Types:
    - ID: Domicile ID (numeric)
    - Date: Domicile date in YYYY-MM-DD format
    - Status: Record status
    - CNIC: CNIC number (13 digits without dashes)
    - Name: First name of applicant
    - Father Name: Father's name of applicant
    - Contact: Contact number
    - Date of Birth: Date in YYYY-MM-DD format
    - Receipt No: Receipt number
    
    If search_type and search_input are not provided, returns all records.
    """
    
    # If search parameters are not provided, return all records
    # if not search_type or not search_input or len(search_input.strip()) == 0:
    SEARCH_MAP = {
        "CNIC": ("d.CNIC LIKE %s", lambda v: f"%{v}%"),
        "Name": ("d.First_Name LIKE %s", lambda v: f"%{v}%"),
        "Father Name": ("d.Father_Name LIKE %s", lambda v: f"%{v}%"),
        "Receipt No": ("d.receipt_no LIKE %s", lambda v: f"%{v}%"),
        "Status": ("d.Status LIKE %s", lambda v: f"%{v}%"),
    }
    try:
        # Base query - select all columns from domicile
        base_query = """
            SELECT d.Dom_id, d.Dom_date, d.Status, d.CNIC, d.First_Name, d.Last_Name,
                   d.Father_Name, d.Spouse_Name, d.Pres_Tehsil, d.Pres_District, d.Pres_Province,
                   d.Present_Address, d.Perm_Tehsil, d.Perm_District, d.Perm_Province, d.Permenant_Address,
                   d.Placeofbirth, d.Contact, d.Date_of_Birth, d.Arrival_Date, d.Gender, d.Religon,
                   d.Marital_Status, d.Qualification, d.Occupation, d.Application_Type, d.Request_Type,
                   d.Service_Type, d.Payment_Type, d.Have_Childern, d.cnic_front, d.cnic_back,
                   d.cnic_guardian, d.Residance_Prof, d.utility_bill, d.educational_certificate,
                   d.marriage_registration_certificate, d.form_b, d.domicile_of_guardian,
                   d.noc_from_concerned_district, d.affidavit_domicile, d.affidavit_voterlist,
                   d.voter_list, d.domicile_challan, d.Approver_Desig, d.Process_Type, d.Purpuse,
                   d.user_id, d.created_at, d.approver_id, d.receipt_no, d.remarks, d.noc_letter_id,
                   di.Dis_Name
            FROM domicile as d 
            JOIN districts as di on d.Perm_District = di.ID
        """
        
        # Build query based on search type
        params = []

        if search_type and search_input:
            if search_type not in SEARCH_MAP:
                raise HTTPException(status_code=400, detail="Invalid search type")

            clause, formatter = SEARCH_MAP[search_type]
            base_query += f" WHERE {clause}"
            params.append(formatter(search_input))

        base_query += " ORDER BY d.Dom_id DESC LIMIT 200"

        con, cur = open_con()
        cur.execute(base_query, params)
        records = cur.fetchall()
        cur.close()
        con.close()

        
        if records:
            # Format records for response - include all columns
            formatted_records = []
            for row in records:
                formatted_records.append({
                    "dom_id": row['Dom_id'],
                    "dom_date": row['Dom_date'],
                    "status": row['Status'],
                    "cnic": row['CNIC'],
                    "first_name": row['First_Name'],
                    "last_name": row['Last_Name'],
                    "father_name": row['Father_Name'],
                    "spouse_name": row['Spouse_Name'],
                    "pres_tehsil": row['Pres_Tehsil'],
                    "pres_district": row['Pres_District'],
                    "pres_province": row['Pres_Province'],
                    "present_address": row['Present_Address'],
                    "perm_tehsil": row['Perm_Tehsil'],
                    "perm_district": row['Perm_District'],
                    "perm_province": row['Perm_Province'],
                    "permenant_address": row['Permenant_Address'],
                    "placeofbirth": row['Placeofbirth'],
                    "contact": row['Contact'],
                    "date_of_birth": row['Date_of_Birth'],
                    "arrival_date": row['Arrival_Date'],
                    "gender": row['Gender'],
                    "religon": row['Religon'],
                    "marital_status": row['Marital_Status'],
                    "qualification": row['Qualification'],
                    "occupation": row['Occupation'],
                    "application_type": row['Application_Type'],
                    "request_type": row['Request_Type'],
                    "service_type": row['Service_Type'],
                    "payment_type": row['Payment_Type'],
                    "have_childern": row['Have_Childern'],
                    "cnic_front": row['cnic_front'],
                    "cnic_back": row['cnic_back'],
                    "cnic_guardian": row['cnic_guardian'],
                    "residance_prof": row['Residance_Prof'],
                    "utility_bill": row['utility_bill'],
                    "educational_certificate": row['educational_certificate'],
                    "marriage_registration_certificate": row['marriage_registration_certificate'],
                    "form_b": row['form_b'],
                    "domicile_of_guardian": row['domicile_of_guardian'],
                    "noc_from_concerned_district": row['noc_from_concerned_district'],
                    "affidavit_domicile": row['affidavit_domicile'],
                    "affidavit_voterlist": row['affidavit_voterlist'],
                    "voter_list": row['voter_list'],
                    "domicile_challan": row['domicile_challan'],
                    "approver_desig": row['Approver_Desig'],
                    "process_type": row['Process_Type'],
                    "purpuse": row['Purpuse'],
                    "user_id": row['user_id'],
                    "created_at": str(row['created_at']) if row['created_at'] else None,
                    "approver_id": row['approver_id'],
                    "receipt_no": row['receipt_no'],
                    "remarks": row['remarks'],
                    "noc_letter_id": row['noc_letter_id'],
                    "district": row['Dis_Name']
                })
            return {"records": formatted_records, "count": len(formatted_records)}
        else:
            return {"records": [], "count": 0, "message": "No records found"}
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error searching records for search_type=%s, search_input=%s", search_type, search_input)
        raise HTTPException(status_code=500, detail=f"Something went wrong: {str(exc)}")


@router.post("/store")
def store_domicile(data: dict):
    """
    Store a new domicile record with support for multiple children.
    
    Expected data format:
    {
        "domicile": {
            "cnic": "...",
            "first_name": "...",
            "father_name": "...",
            ... (all domicile fields)
        },
        "children": [
            {
                "child_name": "...",
                "child_dob": "YYYY-MM-DD" (optional)
            },
            ...
        ]
    }
    """
    try:
        if not data or "domicile" not in data:
            raise HTTPException(status_code=400, detail="Missing 'domicile' data in request")
        
        domicile_data = data.get("domicile")
        children_data = data.get("children", [])
        
        # Validate required domicile fields
        required_fields = ["cnic", "first_name", "father_name"]
        for field in required_fields:
            if not domicile_data.get(field):
                raise HTTPException(status_code=400, detail=f"Missing required field: {field}")
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        # Check if CNIC already exists
        cur.execute("SELECT dom_id FROM domicile WHERE cnic = %s", [domicile_data["cnic"]])
        existing = cur.fetchone()
        if existing:
            cur.close()
            con.close()
            raise HTTPException(status_code=400, detail="CNIC already exists in the system")
        
        # Insert domicile record
        domicile_query = """
            INSERT INTO Domicile (
                Dom_date, Status, CNIC, First_Name, Father_Name, Spouse_Name, 
                Pres_Tehsil, Pres_District, Pres_Province, Present_Address, 
                Perm_Tehsil, Perm_District, Perm_Province, Permenant_Address, 
                Placeofbirth, Contact, Date_of_Birth, Arrival_Date, Gender, Religon, 
                Marital_Status, Qualification, Occupation, Application_Type, Request_Type, 
                Service_Type, Payment_Type, cnic_front, cnic_back, cnic_guardian, 
                Residance_Prof, utility_bill, educational_certificate, 
                marriage_registration_certificate, form_b, domicile_of_guardian, 
                noc_from_concerned_district, affidavit_domicile, affidavit_voterlist, 
                voter_list, domicile_challan, Process_Type, approver_id, user_id, 
                receipt_no, remarks, Have_Childern, noc_letter_id
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s
            )
        """
        
        # Extract parameters with default values
        params = [
            date.today(),  # Dom_date
            domicile_data.get("status", "Pending"),
            domicile_data.get("cnic"),
            domicile_data.get("first_name"),
            domicile_data.get("father_name"),
            domicile_data.get("spouse_name"),
            domicile_data.get("pres_tehsil"),
            domicile_data.get("pres_district"),
            domicile_data.get("pres_province"),
            domicile_data.get("present_address"),
            domicile_data.get("perm_tehsil"),
            domicile_data.get("perm_district"),
            domicile_data.get("perm_province"),
            domicile_data.get("permenant_address"),
            domicile_data.get("placeofbirth"),
            domicile_data.get("contact"),
            domicile_data.get("date_of_birth"),
            domicile_data.get("arrival_date"),
            domicile_data.get("gender"),
            domicile_data.get("religon"),
            domicile_data.get("marital_status"),
            domicile_data.get("qualification"),
            domicile_data.get("occupation"),
            domicile_data.get("application_type"),
            domicile_data.get("request_type"),
            domicile_data.get("service_type"),
            domicile_data.get("payment_type"),
            domicile_data.get("cnic_front"),
            domicile_data.get("cnic_back"),
            domicile_data.get("cnic_guardian"),
            domicile_data.get("residance_prof"),
            domicile_data.get("utility_bill"),
            domicile_data.get("educational_certificate"),
            domicile_data.get("marriage_registration_certificate"),
            domicile_data.get("form_b"),
            domicile_data.get("domicile_of_guardian"),
            domicile_data.get("noc_from_concerned_district"),
            domicile_data.get("affidavit_domicile"),
            domicile_data.get("affidavit_voterlist"),
            domicile_data.get("voter_list"),
            domicile_data.get("domicile_challan"),
            domicile_data.get("process_type"),
            domicile_data.get("approver_desig"),
            domicile_data.get("user_id"),
            domicile_data.get("receipt_no"),
            domicile_data.get("remarks"),
            1 if children_data else 0,  # Have_Childern
            domicile_data.get("noc_letter_id")
        ]
        
        cur.execute(domicile_query, params)
        con.commit()
        
        # Get the last inserted domicile ID
        cur.execute("SELECT LAST_INSERT_ID() as dom_id")
        result = cur.fetchone()
        new_dom_id = result.get("dom_id") if isinstance(result, dict) else result[0]
        
        # Insert children records if provided
        if children_data and len(children_data) > 0:
            for child in children_data:
                child_name = child.get("child_name", "").strip()
                child_dob = child.get("child_dob")
                
                if not child_name:
                    logger.warning("Skipping child record with empty name for dom_id=%s", new_dom_id)
                    continue
                
                child_query = "INSERT INTO Childern (Father_CNIC, Child_Name, Child_dob) VALUES (%s, %s, %s)"
                child_params = [domicile_data["cnic"], child_name, child_dob if child_dob else None]
                
                try:
                    cur.execute(child_query, child_params)
                    con.commit()
                    logger.info("Saved child: %s for dom_id=%s", child_name, new_dom_id)
                except Exception as e:
                    logger.error("Error saving child %s: %s", child_name, str(e))
                    con.rollback()
                    # Continue processing other children
        
        # Insert history record
        history_query = "INSERT INTO dataentry_history (rec_id, user_id, his_status, timestamp) VALUES (%s, %s, 'Created', CURRENT_TIMESTAMP)"
        cur.execute(history_query, [new_dom_id, domicile_data.get("user_id")])
        con.commit()
        
        cur.close()
        con.close()
        
        return {
            "status": "success",
            "message": "Domicile record created successfully",
            "dom_id": new_dom_id,
            "children_count": len(children_data)
        }
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error storing domicile record")
        raise HTTPException(status_code=500, detail=f"Failed to store domicile record: {str(exc)}")


@router.put("/update/{dom_id}")
def update_domicile(dom_id: int, data: dict):
    """
    Update an existing domicile record with support for children management.
    
    Expected data format:
    {
        "domicile": {
            "cnic": "...",
            "first_name": "...",
            ... (all domicile fields to update)
        },
        "children": {
            "new": [
                {
                    "child_name": "...",
                    "child_dob": "YYYY-MM-DD" (optional)
                }
            ],
            "updated": [
                {
                    "child_id": ...,
                    "child_name": "...",
                    "child_dob": "YYYY-MM-DD" (optional)
                }
            ],
            "deleted": [child_id, child_id, ...]
        }
    }
    """
    try:
        if not data or "domicile" not in data:
            raise HTTPException(status_code=400, detail="Missing 'domicile' data in request")
        
        domicile_data = data.get("domicile")
        children_ops = data.get("children", {})
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        # Check if domicile record exists
        cur.execute("SELECT dom_id FROM domicile WHERE dom_id = %s", [dom_id])
        existing = cur.fetchone()
        if not existing:
            cur.close()
            con.close()
            raise HTTPException(status_code=404, detail=f"Domicile record with ID {dom_id} not found")
        
        # Update domicile record
        update_fields = []
        update_params = []
        
        # Map of data fields to database columns
        field_mapping = {
            "cnic": "CNIC",
            "first_name": "First_Name",
            "father_name": "Father_Name",
            "spouse_name": "Spouse_Name",
            "pres_tehsil": "Pres_Tehsil",
            "pres_district": "Pres_District",
            "pres_province": "Pres_Province",
            "present_address": "Present_Address",
            "perm_tehsil": "Perm_Tehsil",
            "perm_district": "Perm_District",
            "perm_province": "Perm_Province",
            "permenant_address": "Permenant_Address",
            "placeofbirth": "Placeofbirth",
            "contact": "Contact",
            "date_of_birth": "Date_of_Birth",
            "arrival_date": "Arrival_Date",
            "gender": "Gender",
            "religon": "Religon",
            "marital_status": "Marital_Status",
            "qualification": "Qualification",
            "occupation": "Occupation",
            "application_type": "Application_Type",
            "request_type": "Request_Type",
            "service_type": "Service_Type",
            "payment_type": "Payment_Type",
            "cnic_front": "cnic_front",
            "cnic_back": "cnic_back",
            "cnic_guardian": "cnic_guardian",
            "residance_prof": "Residance_Prof",
            "utility_bill": "utility_bill",
            "educational_certificate": "educational_certificate",
            "marriage_registration_certificate": "marriage_registration_certificate",
            "form_b": "form_b",
            "domicile_of_guardian": "domicile_of_guardian",
            "noc_from_concerned_district": "noc_from_concerned_district",
            "affidavit_domicile": "affidavit_domicile",
            "affidavit_voterlist": "affidavit_voterlist",
            "voter_list": "voter_list",
            "domicile_challan": "domicile_challan",
            "status": "Status",
            "process_type": "Process_Type",
            "approver_desig": "Approver_Desig",
            "receipt_no": "receipt_no",
            "remarks": "remarks"
        }
        
        # Build dynamic UPDATE clause
        for key, db_column in field_mapping.items():
            if key in domicile_data:
                update_fields.append(f"{db_column} = %s")
                update_params.append(domicile_data[key])
        
        if update_fields:
            update_query = f"UPDATE domicile SET {', '.join(update_fields)} WHERE dom_id = %s"
            update_params.append(dom_id)
            cur.execute(update_query, update_params)
            con.commit()
        
        # Handle children operations
        children_to_delete = children_ops.get("deleted", [])
        children_to_add = children_ops.get("new", [])
        children_to_update = children_ops.get("updated", [])
        
        # Delete children
        if children_to_delete and len(children_to_delete) > 0:
            for child_id in children_to_delete:
                delete_query = "DELETE FROM Childern WHERE child_id = %s"
                try:
                    cur.execute(delete_query, [child_id])
                    con.commit()
                    logger.info("Deleted child with ID: %s", child_id)
                except Exception as e:
                    logger.error("Error deleting child ID %s: %s", child_id, str(e))
                    con.rollback()
        
        # Update existing children
        if children_to_update and len(children_to_update) > 0:
            for child in children_to_update:
                child_id = child.get("child_id")
                child_name = child.get("child_name", "").strip()
                child_dob = child.get("child_dob")
                
                if not child_id or not child_name:
                    logger.warning("Skipping child update: missing child_id or child_name")
                    continue
                
                update_child_query = "UPDATE Childern SET Child_Name = %s, Child_dob = %s WHERE child_id = %s"
                try:
                    cur.execute(update_child_query, [child_name, child_dob if child_dob else None, child_id])
                    con.commit()
                    logger.info("Updated child ID: %s", child_id)
                except Exception as e:
                    logger.error("Error updating child ID %s: %s", child_id, str(e))
                    con.rollback()
        
        # Add new children
        if children_to_add and len(children_to_add) > 0:
            for child in children_to_add:
                child_name = child.get("child_name", "").strip()
                child_dob = child.get("child_dob")
                
                if not child_name:
                    logger.warning("Skipping child record with empty name for dom_id=%s", dom_id)
                    continue
                
                # Get CNIC from the domicile record
                father_cnic = domicile_data.get("cnic")
                if not father_cnic:
                    # Fetch CNIC from existing record
                    cur.execute("SELECT CNIC FROM domicile WHERE dom_id = %s", [dom_id])
                    cnic_result = cur.fetchone()
                    father_cnic = cnic_result["CNIC"] if isinstance(cnic_result, dict) else cnic_result[0]
                
                add_child_query = "INSERT INTO Childern (Father_CNIC, Child_Name, Child_dob) VALUES (%s, %s, %s)"
                try:
                    cur.execute(add_child_query, [father_cnic, child_name, child_dob if child_dob else None])
                    con.commit()
                    logger.info("Added new child: %s for dom_id=%s", child_name, dom_id)
                except Exception as e:
                    logger.error("Error adding child %s: %s", child_name, str(e))
                    con.rollback()
        
        # Insert history record
        history_query = "INSERT INTO dataentry_history (rec_id, user_id, his_status, timestamp) VALUES (%s, %s, 'Updated', CURRENT_TIMESTAMP)"
        cur.execute(history_query, [dom_id, domicile_data.get("user_id")])
        con.commit()
        
        cur.close()
        con.close()
        
        return {
            "status": "success",
            "message": "Domicile record updated successfully",
            "dom_id": dom_id,
            "children_deleted": len(children_to_delete),
            "children_updated": len(children_to_update),
            "children_added": len(children_to_add)
        }
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error updating domicile record with ID %s", dom_id)
        raise HTTPException(status_code=500, detail=f"Failed to update domicile record: {str(exc)}")


@router.get("/check-duplicate-cnic/{cnic}")
def check_duplicate_cnic(cnic: str):
    """
    Check if a CNIC already exists in the domicile records.
    Returns status indicating if CNIC exists or is available.
    """
    try:
        if not cnic or len(cnic.strip()) == 0:
            raise HTTPException(status_code=400, detail="CNIC is required")
        
        # Clean CNIC - remove non-digits
        clean_cnic = re.sub(r"\D", "", cnic.strip())
        if len(clean_cnic) != 13:
            raise HTTPException(status_code=400, detail="CNIC must be 13 digits")
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        cur.execute("SELECT dom_id FROM domicile WHERE cnic = %s", [clean_cnic])
        result = cur.fetchone()
        
        if result:
            return {
                "status": "exists",
                "action": "stop",
                "message": f"CNIC {clean_cnic} already exists in the system"
            }
        Query = "Select CNIC from cancellation where CNIC = %s;"
        cur.execute(Query, [clean_cnic])
        data = cur.fetchall()
        if data:
            return {
                "status": "exists",
                "action": "continue",
                "message": f"Applicant domicile has been cancelled for this CNIC {clean_cnic}"
            }
        
        Query = "Select cnic from noc_applicants where CNIC = %s;"
        cur.execute(Query, [clean_cnic])
        data = cur.fetchall()
        if data:
           return {
                "status": "exists",
                "action": "stop",
                "message": f"NOC already issued for this CNIC {clean_cnic}"
            }
        Query = "Select CNIC from noc_ict_applicants where CNIC = %s;"
        cur.execute(Query, [clean_cnic])
        data = cur.fetchall()
        if data:
            return {
                "status": "exists",
                "action": "continue",
                "message": f"NOC letter for ICT domicile issued for this CNIC {clean_cnic}"
            }
        Query = "Select cnic, reason from black_list where cnic = %s and status = 'Blocked';"
        cur.execute(Query, [clean_cnic])
        data = cur.fetchall()
        if data:
            return {
                "status": "exists",
                "action": "stop",
                "message": f"This cnic no is black listed due to following reason:-\n {data[0][1]}"
            }
            
        cur.close()
        con.close()   
        return {
            "status": "available",
            "action": "continue",
            "message": f"CNIC {clean_cnic} is available for new domicile application"
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error checking CNIC: %s", cnic)
        raise HTTPException(status_code=500, detail=f"Error checking CNIC: {str(exc)}")


@router.post("/history")
def store_history(rec_id: int = None, user_id: int = None, status: str = "Created"):
    """
    Store a history record for domicile operations.
    
    Parameters:
    - rec_id: Domicile record ID
    - user_id: User ID performing the operation
    - status: Status of the operation (e.g., 'Created', 'Updated', 'Approved')
    """
    try:
        if not rec_id:
            raise HTTPException(status_code=400, detail="rec_id is required")
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        # Verify domicile record exists
        cur.execute("SELECT dom_id FROM domicile WHERE dom_id = %s", [rec_id])
        dom_check = cur.fetchone()
        if not dom_check:
            cur.close()
            con.close()
            raise HTTPException(status_code=404, detail=f"Domicile record {rec_id} not found")
        
        # Insert history record
        history_query = """
            INSERT INTO dataentry_history (rec_id, user_id, his_status, timestamp)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        """
        cur.execute(history_query, [rec_id, user_id, status])
        con.commit()
        
        cur.execute("SELECT LAST_INSERT_ID() as history_id")
        history_result = cur.fetchone()
        history_id = history_result.get("history_id") if isinstance(history_result, dict) else history_result[0]
        
        cur.close()
        con.close()
        
        logger.info("History record created for rec_id=%s, user_id=%s, status=%s", rec_id, user_id, status)
        
        return {
            "status": "success",
            "message": "History record stored successfully",
            "history_id": history_id,
            "rec_id": rec_id,
            "user_id": user_id,
            "operation": status
        }
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error storing history record for rec_id=%s", rec_id)
        raise HTTPException(status_code=500, detail=f"Failed to store history: {str(exc)}")


@router.get("/history/{rec_id}")
def get_history(rec_id: int):
    """
    Retrieve all history records for a specific domicile record.
    """
    try:
        if not rec_id:
            raise HTTPException(status_code=400, detail="rec_id is required")
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        # Verify domicile record exists
        cur.execute("SELECT dom_id FROM domicile WHERE dom_id = %s", [rec_id])
        dom_check = cur.fetchone()
        if not dom_check:
            cur.close()
            con.close()
            raise HTTPException(status_code=404, detail=f"Domicile record {rec_id} not found")
        
        # Fetch history
        history_query = """
            SELECT h.his_status, h.timestamp, u.user_name 
            FROM dataentry_history as h 
            LEFT JOIN users as u ON h.user_id = u.user_id 
            WHERE h.rec_id = %s
            ORDER BY h.timestamp DESC
        """
        cur.execute(history_query, [rec_id])
        history_records = cur.fetchall()
        cur.close()
        con.close()
        
        formatted_history = []
        if history_records:
            for row in history_records:
                formatted_history.append({
                    "status": row['his_status'],
                    "timestamp": str(row['timestamp']) if row['timestamp'] else None,
                    "user_name": row['user_name'] if row['user_name'] else "Unknown"
                })
        
        return {
            "rec_id": rec_id,
            "history": formatted_history,
            "total_records": len(formatted_history)
        }
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error retrieving history for rec_id=%s", rec_id)
        raise HTTPException(status_code=500, detail=f"Failed to retrieve history: {str(exc)}")


@router.post("/history/receipt-change")
def store_receipt_change_history(rec_id: int, old_receipt_no: str, new_receipt_no: str, user_id: int):
    """
    Store a history record when receipt number is changed.
    """
    try:
        if not rec_id:
            raise HTTPException(status_code=400, detail="rec_id is required")
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        # Verify domicile record exists
        cur.execute("SELECT dom_id FROM domicile WHERE dom_id = %s", [rec_id])
        dom_check = cur.fetchone()
        if not dom_check:
            cur.close()
            con.close()
            raise HTTPException(status_code=404, detail=f"Domicile record {rec_id} not found")
        
        # Create history message
        status_message = f"Receipt No changed from {old_receipt_no} to {new_receipt_no}"
        
        # Insert history record
        history_query = """
            INSERT INTO dataentry_history (rec_id, user_id, his_status, timestamp)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
        """
        cur.execute(history_query, [rec_id, user_id, status_message])
        con.commit()
        
        cur.close()
        con.close()
        
        logger.info("Receipt change history recorded for rec_id=%s", rec_id)
        
        return {
            "status": "success",
            "message": "Receipt change history recorded",
            "rec_id": rec_id,
            "old_receipt_no": old_receipt_no,
            "new_receipt_no": new_receipt_no
        }
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error storing receipt change history for rec_id=%s", rec_id)
        raise HTTPException(status_code=500, detail=f"Failed to store receipt change history: {str(exc)}")


@router.post("/pictures")
def store_picture(data: dict):
    """
    Store a picture blob in the pictures table.
    
    Expected data format:
    {
        "dom_id": <integer>,
        "pic": "<base64_encoded_image>",
        "ext": "<file_extension>" (optional, e.g., "jpg", "png")
    }
    """
    try:
        if not data:
            raise HTTPException(status_code=400, detail="Missing request data")
        
        dom_id = data.get("dom_id")
        pic_b64 = data.get("pic")
        ext = data.get("ext")
        
        # Validate required fields
        if not dom_id:
            raise HTTPException(status_code=400, detail="Missing required field: dom_id")
        if not pic_b64:
            raise HTTPException(status_code=400, detail="Missing required field: pic")
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        try:
            # Verify domicile record exists and get CNIC
            cur.execute("SELECT cnic FROM domicile WHERE dom_id = %s", [dom_id])
            domicile = cur.fetchone()
            if not domicile:
                cur.close()
                con.close()
                raise HTTPException(status_code=404, detail=f"Domicile record with dom_id {dom_id} not found")
            
            cnic = domicile['cnic']
            print(domicile)
            # Decode base64 picture to binary
            try:
                pic_binary = base64.b64decode(pic_b64)
            except Exception as e:
                logger.error("Failed to decode base64 picture: %s", str(e))
                raise HTTPException(status_code=400, detail="Invalid base64 encoded picture")
            
            # Insert picture record
            picture_query = """
                INSERT INTO pictures (cnic, Pic, ext, data_validity, Dom_id)
                VALUES (%s, %s, %s, %s, %s)
            """
            cur.execute(picture_query, [cnic, pic_binary, ext, 0, dom_id])
            con.commit()
            
            pic_id = cur.lastrowid
            cur.close()
            con.close()
            
            logger.info("Picture stored successfully for dom_id=%s, pic_id=%s", dom_id, pic_id)
            
            return {
                "status": "success",
                "message": "Picture stored successfully",
                "pic_id": pic_id,
                "dom_id": dom_id,
                "cnic": cnic
            }
        
        except HTTPException:
            raise
        except Exception as e:
            cur.close()
            con.close()
            logger.exception("Error storing picture for dom_id=%s", dom_id)
            raise HTTPException(status_code=500, detail=f"Failed to store picture: {str(e)}")
    
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error in store_picture endpoint")
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(exc)}")


# ============== CANCELLATION LETTER ROUTES ==============

@router.get('/cancellations/search')
def search_cancellations(
    keyword: str = Query(None, description="Search field: CNIC, Name, Father Name, Domicile No, Dispatch No, Domicile Date, Letter Date"),
    value: str = Query(None, description="Search value"),
    query_type: str = Query("None", description="'All' for all records, 'None' for filtered")):
    """Search cancellation letters with various filters."""
    try:
        if not value or value == 'None' or len(value.strip()) == 0:
            query_type = 'All'
            query_part = ""
        else:
            query_part = ""
            if keyword == "CNIC":
                query_part = f"c.CNIC like '%{value}%'"
            elif keyword == "Name":
                query_part = f"c.Applicant_Name like '%{value}%'"
            elif keyword == "Father Name":
                query_part = f"c.Father_Name like '%{value}%'"
            elif keyword == "Domicile No":
                query_part = f"c.Domicile_No like '%{value}%'"
            elif keyword == "Dispatch No":
                query_part = f"d.Dispatch_No = '{value}'"
            elif keyword == "Domicile Date":
                if ' ' in value:
                    val1, val2 = value.split(' ', 1)
                    query_part = f"c.Domicile_Date Between '{val1}' and '{val2}'"
                else:
                    query_part = f"c.Domicile_Date = '{value}'"
            elif keyword == "Letter Date":
                if ' ' in value:
                    val1, val2 = value.split(' ', 1)
                    query_part = f"c.Letter_Date Between '{val1}' and '{val2}'"
                else:
                    query_part = f"c.Letter_Date = '{value}'"
        
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        base_query = """SELECT 
                c.Letter_ID,
                d.Dispatch_No,
                c.Letter_Date,
                c.CNIC,
                c.Applicant_Name,
                c.Relation,
                c.Father_Name,
                c.Address,
                c.Domicile_No,
                c.Domicile_Date
            FROM cancellation as c
            INNER JOIN dispatch_dairy as d
                ON c.Letter_ID = d.Letter_ID and d.Letter_Type = 'Cancellation Letter' """
        
        if query_type == 'All':
            full_query = base_query + "ORDER BY c.Letter_ID DESC LIMIT 200;"
        else:
            if not query_part:
                full_query = base_query + "ORDER BY c.Letter_ID DESC LIMIT 200;"
            else:
                full_query = base_query + f" WHERE {query_part} ORDER BY c.Letter_ID DESC;"
        
        cur.execute(full_query)
        rows = cur.fetchall()
        cur.close()
        con.close()
        
        return {"cancellations": rows}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to search cancellation letters")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/cancellations/{letter_id}')
def get_cancellation(letter_id: int):
    """Get detailed cancellation letter data by ID."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        query = "SELECT * FROM cancellation WHERE Letter_ID = %s;"
        cur.execute(query, [letter_id])
        data = cur.fetchone()
        cur.close()
        con.close()
        
        if not data:
            raise HTTPException(status_code=404, detail=f"Cancellation letter {letter_id} not found")
        
        return {"cancellation": data}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch cancellation letter %s", letter_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post('/cancellations')
def create_cancellation(
    cnic: str = Query(..., description="CNIC number"),
    applicant_name: str = Query(..., description="Applicant name"),
    relation: str = Query(..., description="Relation: s/o, d/o, w/o"),
    father_name: str = Query(..., description="Father/Husband name"),
    address: str = Query(..., description="Address"),
    domicile_no: str = Query(..., description="Domicile number"),
    domicile_date: str = Query(..., description="Domicile date (YYYY-MM-DD)"),
    remarks: str = Query("", description="Optional remarks")):
    """Create a new cancellation letter record."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        # Insert into cancellation table
        insert_query = """INSERT INTO cancellation 
                         (CNIC, Applicant_Name, Relation, Father_name, Address, Domicile_No, Domicile_Date, Remarks) 
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"""
        params = [cnic, applicant_name, relation, father_name, address, domicile_no, domicile_date, remarks]
        cur.execute(insert_query, params)
        con.commit()
        
        # Get last inserted ID
        cur.execute('SELECT LAST_INSERT_ID();')
        last_letter_id = cur.fetchone()[0]
        
        # Get next dispatch number
        dispatch_query = "SELECT Dispatch_No, YEAR(timestamp) as Last_Year, YEAR(CURDATE()) as Cur_Year FROM dispatch_dairy ORDER BY Dispatch_ID DESC LIMIT 1;"
        cur.execute(dispatch_query)
        last_dispatch_data = cur.fetchone()
        
        if last_dispatch_data is not None:
            if last_dispatch_data[1] != last_dispatch_data[2]:
                next_dispatch = 1
            else:
                next_dispatch = 1 + last_dispatch_data[0]
        else:
            next_dispatch = 1
        
        # Insert into dispatch_dairy table
        dispatch_insert = "INSERT INTO dispatch_dairy (Dispatch_No, Letter_Type, Letter_ID) VALUES (%s, 'Cancellation Letter', %s);"
        cur.execute(dispatch_insert, [next_dispatch, last_letter_id])
        con.commit()
        cur.close()
        con.close()
        
        return {
            "status": "success",
            "message": "Cancellation letter created",
            "letter_id": last_letter_id,
            "dispatch_no": next_dispatch
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to create cancellation letter")
        raise HTTPException(status_code=500, detail=str(exc))


@router.put('/cancellations/{letter_id}')
def update_cancellation(
    letter_id: int,
    cnic: str = Query(..., description="CNIC number"),
    applicant_name: str = Query(..., description="Applicant name"),
    relation: str = Query(..., description="Relation: s/o, d/o, w/o"),
    father_name: str = Query(..., description="Father/Husband name"),
    address: str = Query(..., description="Address"),
    domicile_no: str = Query(..., description="Domicile number"),
    domicile_date: str = Query(..., description="Domicile date (YYYY-MM-DD)"),
    remarks: str = Query("", description="Optional remarks")):
    """Update an existing cancellation letter record."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        update_query = """UPDATE cancellation 
                         SET CNIC = %s, Applicant_Name = %s, Father_Name = %s, Relation = %s, 
                             Address = %s, Domicile_No = %s, Domicile_Date = %s, Remarks = %s
                         WHERE Letter_ID = %s;"""
        params = [cnic, applicant_name, father_name, relation, address, domicile_no, domicile_date, remarks, letter_id]
        cur.execute(update_query, params)
        con.commit()
        
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"Cancellation letter {letter_id} not found")
        
        cur.close()
        con.close()
        
        return {
            "status": "success",
            "message": "Cancellation letter updated",
            "letter_id": letter_id
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to update cancellation letter %s", letter_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/cancellations/{letter_id}/print')
def get_cancellation_for_print(letter_id: int):
    """Get cancellation letter data formatted for printing/PDF generation."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        query = """SELECT a.Letter_ID, d.Dispatch_No, a.Letter_Date, a.CNIC, a.Applicant_Name, a.Relation, 
                          a.Father_Name, a.Address, a.Domicile_No, a.Domicile_Date 
                   FROM cancellation as a                                
                   INNER JOIN dispatch_dairy as d
                       ON d.Letter_ID = a.Letter_ID
                   WHERE a.Letter_ID = %s AND d.Letter_Type = 'Cancellation Letter';"""
        cur.execute(query, [letter_id])
        data = cur.fetchone()
        cur.close()
        con.close()
        
        if not data:
            raise HTTPException(status_code=404, detail=f"Cancellation letter {letter_id} not found")
        
        return {"print_data": data}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch cancellation letter print data for %s", letter_id)
        raise HTTPException(status_code=500, detail=str(exc))


# ============== CASH REPORT ROUTES ==============

@router.post("/cash-reports/search")
def search_cash_reports(filters: dict = Body(None)):
    
    CASH_REPORT_FILTERS = {
        "CNIC": "CNIC",
        "Payment_Type": "Payment_Type",
        "Application_Type": "Application_Type",
        "Request_Type": "Request_Type",
        "Duplicate_Entry": "Duplicate_Entry",
        "Domicile_Date": "Domicile_Date"
    }

    try:
        con, cur = open_con()
        if type(con) is str:
            raise HTTPException(status_code=500, detail=cur)

        base_query = """
            SELECT Domicile_ID, Domicile_Date, CNIC, Applicant_Name,
                   Payment_Type, Govt_Fee, Application_Type,
                   Request_Type, Duplicate_Entry
            FROM cash_report
        """

        conditions = []
        params = []

        if filters:
            for key, value in filters.items():
                if key not in CASH_REPORT_FILTERS:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Invalid filter field: {key}"
                    )
                conditions.append(f"{CASH_REPORT_FILTERS[key]} = %s")
                params.append(value)

        if conditions:
            base_query += " WHERE " + " AND ".join(conditions)
        print(base_query, params)
        cur.execute(base_query, params)
        rows = cur.fetchall()
        cur.close()
        con.close()

        return {
            "total_records": len(rows),
            "cash_reports": rows
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Cash report search failed")
        raise HTTPException(status_code=500, detail=str(exc))



@router.get('/cash-reports/{domicile_id}')
def get_cash_report(domicile_id: int):
    """Get a specific cash report record by Domicile_ID."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        query = "SELECT * FROM cash_report WHERE Domicile_ID = %s;"
        cur.execute(query, [domicile_id])
        data = cur.fetchone()
        cur.close()
        con.close()
        
        if not data:
            raise HTTPException(status_code=404, detail=f"Cash report {domicile_id} not found")
        
        return {"cash_report": data}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to fetch cash report %s", domicile_id)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post('/cash-reports')
def create_cash_report(
    domicile_date: str = Query(..., description="Domicile date (YYYY-MM-DD)"),
    applicant_name: str = Query(..., description="Applicant name"),
    cnic: str = Query(..., description="CNIC number"),
    payment_type: str = Query(..., description="Payment Type"),
    govt_fee: str = Query(..., description="Govt Fee"),
    application_type: str = Query(..., description="Application Type (offline/online)"),
    duplicate_entry: str = Query(..., description="Duplicate Entry (Yes/No)"),
    domicile_no: str = Query(..., description="Domicile number"),
    request_type: str = Query(..., description="Request Type (New/Revised/Duplicate)")):
    """Create a new cash report record."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        insert_query = """INSERT INTO cash_report 
                         (Domicile_Date, Applicant_Name, cnic, Payment_Type, Govt_Fee, 
                          Application_Type, Duplicate_Entry, Domicile_No, Request_Type) 
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        params = [domicile_date, applicant_name, cnic, payment_type, govt_fee, 
                 application_type, duplicate_entry, domicile_no, request_type]
        cur.execute(insert_query, params)
        con.commit()
        
        # Get last inserted ID
        cur.execute('SELECT LAST_INSERT_ID();')
        last_id = cur.fetchone()[0]
        
        cur.close()
        con.close()
        
        return {
            "status": "success",
            "message": "Cash report record created",
            "domicile_id": last_id
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to create cash report")
        raise HTTPException(status_code=500, detail=str(exc))


@router.get('/cash-reports/check-date/{report_date}')
def check_date_exists(report_date: str):
    """Check if a domicile date already exists in cash_report table."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        query = "SELECT Domicile_Date FROM cash_report WHERE Domicile_Date = %s;"
        cur.execute(query, [report_date])
        data = cur.fetchall()
        cur.close()
        con.close()
        
        if data and len(data) > 0:
            return {"status": "exists", "message": "Date already exist"}
        else:
            return {"status": "not_exists", "message": "Date not exist"}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to check date existence for %s", report_date)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/cash-reports/bulk-insert")
def bulk_insert_cash_reports(records: list = Body(...)):

    """Bulk insert cash report records from Excel."""
    try:
        con, cur = open_con()
        if type(con) is str:
            logger.error("Database connection failed: %s", cur)
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
        
        insert_query = """INSERT INTO cash_report 
                         (Domicile_Date, Applicant_Name, cnic, Application_Type, Duplicate_Entry, 
                          Govt_Fee, Payment_Type, Domicile_No, Request_Type) 
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        
        inserted_count = 0
        for record in records:
            try:
                params = [
                    record.get('domicile_date'),
                    record.get('applicant_name'),
                    record.get('cnic'),
                    record.get('application_type'),
                    record.get('duplicate_entry'),
                    record.get('govt_fee'),
                    record.get('payment_type'),
                    record.get('domicile_no'),
                    record.get('request_type')
                ]
                cur.execute(insert_query, params)
                con.commit()
                inserted_count += 1
            except Exception as record_error:
                logger.warning("Failed to insert record: %s", record_error)
                continue
        
        cur.close()
        con.close()
        
        return {
            "status": "success",
            "message": f"Inserted {inserted_count} records",
            "total_records": len(records),
            "inserted_count": inserted_count
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to bulk insert cash reports")
        raise HTTPException(status_code=500, detail=str(exc))


@router.post('/cash-reports/upload')
def upload_cash_reports(file: UploadFile = File(...)):
    """Accept an Excel file upload, parse and insert cash report records."""
    try:
        contents = file.file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
        if 'Worksheet' not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Worksheet named 'Worksheet' not found in workbook")
        ws = wb['Worksheet']

        records_to_insert = []
        cnic_list = set()
        row_index = 0
        # iterate rows
        for row in ws.values:
            row_index += 1
            if row_index == 1:
                continue
            if row_index == 2:
                try:
                    date_dict = datetime.strptime(row[0], '%d/%m/%Y').strftime('%Y-%m-%d')
                except Exception:
                    raise HTTPException(status_code=400, detail='Invalid date format in file')
                # check date exists
                con, cur = open_con()
                if type(con) is str:
                    raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
                cur.execute("SELECT Domicile_Date FROM cash_report WHERE Domicile_Date = %s;", [date_dict])
                existing = cur.fetchall()
                cur.close()
                con.close()
                if existing and len(existing) > 0:
                    return {"status": "error", "detail": "Date already exist"}

            if row[2] is None:
                continue
            cnic_val = str(row[2]).strip().strip("'")
            dup_value = 'No' if cnic_val not in cnic_list else 'Yes'
            cnic_list.add(cnic_val)

            if row[5] is None:
                Govt_Fee = "0"
                Payment_Type = "Not paid"
                app_type = 'online'
            elif str(row[5]).strip() == 'Cash':
                Govt_Fee = "200"
                Payment_Type = "Cash"
                app_type = 'offline'
            else:
                Govt_Fee = "Paid in Bank"
                Payment_Type = "Challan"
                app_type = 'offline'

            if str(row[6]) == 'Revised' or str(row[6]) == 'Duplicate':
                Govt_Fee = "0"
                Payment_Type = "Not paid"

            try:
                my_date_format = datetime.strptime(row[0], '%d/%m/%Y').strftime('%Y-%m-%d')
            except Exception:
                my_date_format = row[0]

            domicile_no = ''
            try:
                domicile_no = str(row[7])
                idx = domicile_no.find('-', 6)
                if idx != -1:
                    domicile_no = domicile_no[idx+1:]
            except Exception:
                domicile_no = ''

            record = {
                'domicile_date': my_date_format,
                'applicant_name': row[1],
                'cnic': cnic_val,
                'application_type': app_type,
                'duplicate_entry': dup_value,
                'govt_fee': Govt_Fee,
                'payment_type': Payment_Type,
                'domicile_no': domicile_no,
                'request_type': row[6]
            }
            records_to_insert.append(record)

        wb.close()

        # insert into DB
        con, cur = open_con()
        if type(con) is str:
            raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")

        insert_query = """INSERT INTO cash_report 
                         (Domicile_Date, Applicant_Name, cnic, Application_Type, Duplicate_Entry, 
                          Govt_Fee, Payment_Type, Domicile_No, Request_Type) 
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        inserted = 0
        for r in records_to_insert:
            try:
                params = [r['domicile_date'], r['applicant_name'], r['cnic'], r['application_type'], r['duplicate_entry'], r['govt_fee'], r['payment_type'], r['domicile_no'], r['request_type']]
                cur.execute(insert_query, params)
                con.commit()
                inserted += 1
            except Exception:
                continue

        cur.close()
        con.close()

        return {"status": "success", "inserted_count": inserted, "total_records": len(records_to_insert)}

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to process uploaded cash reports")
        raise HTTPException(status_code=500, detail=str(exc))
@router.get('/version')
def version():
    con, cur = open_con()
    if type(con) is str:
        raise HTTPException(status_code=500, detail=f"Database connection failed: {cur}")
    cur.execute('Select ver_no from versions order by ver_id desc limit 1;')
    ver_data = cur.fetchone()
    if ver_data:
        return {"status": "success", "data": ver_data}
    else:
        return {'status':"failed", "data":None, "message":"Version data not available on server."}
    